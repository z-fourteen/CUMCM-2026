# -*- coding: utf-8 -*-
"""问题 4 输出：−100 s ~ 100 s 每秒各把手位置与速度 -> result4.xlsx。

时间与弧长的对应
----------------
龙头前把手速度恒为 1 m/s，调头开始（龙头到达切点 A）为零时刻，
故 s_head = t（单位一致），t ∈ [-100, 100] 对应 s_head ∈ [-100, 100]。

链的初始化必须放在"全队都在盘入螺线上"处（根唯一），再连续 march 到
−100 s，才能保证每个把手锁定在正确的延拓分支上。全队长 369.16 m，
故自 s_head = −(369.16 + 40) 起步。
"""
import sys, os
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from dragon4 import (make, Path, Team, Chain, follow_stats,  # noqa: E402
                     signed_gap_P, CHORDS, BODY, N_BENCH)

HERE = os.path.dirname(os.path.abspath(__file__))
TPL = os.path.join(HERE, '..', '附件', 'result4.xlsx')

T0, T1 = -100.0, 100.0
S_START = -(BODY + 40.0)          # 起步点：全队均在盘入螺线上
DS = 0.05                         # march 步长

# 论文中需列出的把手（0 基）：龙头前、第1/51/101/151/201 节龙身前、龙尾后
KEY = [0, 1, 51, 101, 151, 201, 223]
KEY_NAME = ['龙头前把手', '第1节龙身前把手', '第51节龙身前把手',
            '第101节龙身前把手', '第151节龙身前把手', '第201节龙身前把手',
            '龙尾后把手']


def load_best():
    f = os.path.join(HERE, 'q4_best.npy')
    if os.path.exists(f):
        rA, rB = np.load(f)
        return float(rA), float(rB)
    return 4.279, 4.479


def solve_traj(rA, rB, ts=None, v_head=1.0):
    """-> (g, p, ts, P[nt,224,2], V[nt,224])

    s_head = v_head * t（调头开始为零时刻）。用 Team 查表直接定位，
    无须时间推进：phi' > 0 时 phi 单值，S_i = phi^i(s_head) 唯一。
    """
    g = make(rA, rB)
    assert g is not None, '几何不可行'
    p = Path(g)
    ts = np.arange(T0, T1 + 1e-9, 1.0) if ts is None else np.asarray(ts, float)
    tm = Team(p)
    P = np.empty((len(ts), N_BENCH + 1, 2))
    V = np.empty((len(ts), N_BENCH + 1))
    for k, t in enumerate(ts):
        _, P[k], V[k] = tm.state(v_head * float(t), v_head)
    return g, p, ts, P, V


def write_xlsx(ts, P, V, out):
    import openpyxl
    wb = openpyxl.load_workbook(TPL)
    ws_p, ws_v = wb['位置'], wb['速度']
    # 列 1 是行标签，时间自第 2 列起；模板行序为 x,y 交替（位置）/ 每把手一行（速度）
    for k in range(len(ts)):
        c = 2 + k
        for i in range(N_BENCH + 1):
            ws_p.cell(2 + 2 * i, c).value = round(float(P[k, i, 0]), 6)
            ws_p.cell(3 + 2 * i, c).value = round(float(P[k, i, 1]), 6)
            ws_v.cell(2 + i, c).value = round(float(V[k, i]), 6)
    wb.save(out)


if __name__ == '__main__':
    rA, rB = load_best()
    print('=' * 76)
    print('问题 4 输出：最优调头曲线 rA=%.6f rB=%.6f' % (rA, rB))
    print('=' * 76)
    g, p, ts, P, V = solve_traj(rA, rB)
    print('R1=%.6f R2=%.6f  a1=%.6f a2=%.6f  L=%.6f m'
          % (g['R1'], g['R2'], g['a1'], g['a2'], p.L))
    print('调头曲线跨时段 t ∈ [0, %.4f] s' % p.L)

    print('\n[1] 一致性校核')
    err = 0.0
    for k in range(len(ts)):
        e = np.abs(np.linalg.norm(P[k, :-1] - P[k, 1:], axis=1) - CHORDS).max()
        err = max(err, float(e))
    print('  各时刻板凳弦长最大误差 = %.3e m' % err)
    print('  速度范围 [%.6f, %.6f] m/s   (龙头恒 1)' % (V.min(), V.max()))
    kmax = np.unravel_index(V.argmax(), V.shape)
    print('  最大速度 %.6f m/s 出现在 t=%.0f s 第 %d 把手'
          % (V.max(), ts[kmax[0]], kmax[1]))
    print('  速度是否全程为正: %s' % bool((V > 0).all()))

    print('\n[2] 碰撞校核（每秒采样）')
    Gs = np.array([signed_gap_P(P[k])[0] for k in range(len(ts))])
    k = int(Gs.argmin())
    print('  最小间隙 G = %+.6f m @ t=%.0f s, 板凳对 %s'
          % (Gs[k], ts[k], signed_gap_P(P[k])[1]))

    print('\n[3] 论文所需 5 个时刻 x 7 个把手')
    for t in (-100.0, -50.0, 0.0, 50.0, 100.0):
        k = int(np.argmin(np.abs(ts - t)))
        print('  t = %+.0f s' % t)
        print('    %-16s %10s %10s %10s' % ('把手', 'x (m)', 'y (m)', 'v (m/s)'))
        for i, nm in zip(KEY, KEY_NAME):
            print('    %-16s %10.6f %10.6f %10.6f'
                  % (nm, P[k, i, 0], P[k, i, 1], V[k, i]))

    out = os.path.join(HERE, 'result4.xlsx')
    write_xlsx(ts, P, V, out)
    print('\n已写出 %s' % out)
    np.savez(os.path.join(HERE, 'q4_traj.npz'), ts=ts, P=P, V=V,
             rA=rA, rB=rB, L=p.L)
    print('已保存轨迹 -> q4_traj.npz')
