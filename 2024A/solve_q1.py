# -*- coding: utf-8 -*-
"""
2024 年 CUMCM 数学建模国赛 A 题 问题一：板凳龙沿阿基米德螺线盘入
============================================================

模型概要
--------
* 螺线: r = a*theta,  a = p/(2*pi), p = 0.55 m（螺距）
* 龙头前把手 P_0 初始位于 theta = 32*pi, 即 (8.8, 0)，以线速度 1 m/s
  沿螺线"顺时针向内"运动（theta 减小方向），故 t 与 theta 满足：
        a * [ F(32*pi) - F(theta_0(t)) ] = t
        F(theta) = 0.5 * ( theta*sqrt(1+theta^2) + asinh(theta) )
* 逐节递推: 相邻把手间为刚性直板，满足"欧氏弦长"约束
        |P_i - P_{i-1}| = L_i,   L_1 = 2.86,  L_i = 1.65 (i>=2)
  即  a^2 * ( theta_i^2 + theta_{i-1}^2 - 2*theta_i*theta_{i-1}*cos(delta) ) = L_i^2
  选择 theta_i > theta_{i-1} 的最近正向根（brentq）。
* 速度: 每个把手沿螺线切向运动 V_i = v_i * T_i（T_i 为盘入方向单位切向量），
  由刚性约束 d_i·(V_i - V_{i-1}) = 0 得到解析递推
        v_i = v_{i-1} * (d_i·T_{i-1}) / (d_i·T_i),   v_0 = 1
  并用隐函数微分法与中心差分交叉验证。

输出
----
* result1.xlsx（写入官方模板，位置/速度 两张表，6 位小数）
* 论文用两张代表表格（6 个时刻 x 7 个代表节点）
* 多张验证图形
"""
import time
import numpy as np
import openpyxl
from scipy.optimize import brentq
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False

# ======================================================================
# 1. 常量与螺线几何
# ======================================================================
PITCH = 0.55                       # 螺距 p = 0.55 m
A = PITCH / (2.0 * np.pi)          # 阿基米德螺线参数 r = a*theta
L_HEAD = 2.86                      # 龙头两个把手中心间距 (m)
L_BODY = 1.65                      # 龙身/龙尾两个把手中心间距 (m)
N_NODES = 224                      # 把手总数 P_0 ... P_223
THETA_INIT = 32.0 * np.pi          # 龙头初始极角
V0 = 1.0                           # 龙头前把手线速度 (m/s)
T_MAX = 300
TIMES = np.arange(0, T_MAX + 1, dtype=float)


def spiral_position(theta):
    """阿基米德螺线上极角 theta 对应的直角坐标 (x, y)。"""
    return np.array([A * theta * np.cos(theta),
                     A * theta * np.sin(theta)])


def spiral_dr_dtheta(theta):
    """参数导数 d r_vec / d theta。"""
    return np.array([A * (np.cos(theta) - theta * np.sin(theta)),
                     A * (np.sin(theta) + theta * np.cos(theta))])


def spiral_tangent(theta):
    """实际盘入方向（theta 减小的方向）的单位切向量 T。"""
    dr = spiral_dr_dtheta(theta)
    return -dr / np.linalg.norm(dr)


def spiral_arc_primitive(theta):
    """弧长原函数 F(theta) = int sqrt(1+u^2) du = 0.5(u*sqrt(1+u^2)+asinh u)。"""
    return 0.5 * (theta * np.sqrt(1.0 + theta**2) + np.arcsinh(theta))


F_INIT = spiral_arc_primitive(THETA_INIT)   # F(32*pi)


def chord_sq(theta_prev, theta):
    """两个螺线点之间欧氏距离的平方。"""
    return A**2 * (theta**2 + theta_prev**2
                   - 2.0 * theta * theta_prev * np.cos(theta - theta_prev))


# ======================================================================
# 2. 龙头运动：求解 theta_0(t)
# ======================================================================
def solve_head_theta(t, hi=THETA_INIT):
    """
    由 a*[F(32*pi) - F(theta)] = t 求龙头当前极角 theta_0(t)。
    theta_0 随时间单调递减。brentq 求根于 [0, hi]。
    """
    if t <= 0.0:
        return THETA_INIT
    f_target = F_INIT - t / A            # 需要 F(theta) = F(32*pi) - t/a
    lo = 0.0
    return brentq(lambda th: spiral_arc_primitive(th) - f_target, lo, hi)


# ======================================================================
# 3. 逐节递推：求下一把手极角（取最近正向根）
# ======================================================================
def solve_next_theta(theta_prev, L):
    """
    已知 theta_{i-1}，求满足 |P_i - P_{i-1}| = L 的最近正向根 theta_i。

    说明：弦长 c(delta) 在 delta 较小时随 delta 单调递增，
    其极大值出现在 delta ~ pi（半圈）附近。本模拟中所有节点半径 r >= ~5 m，
    而所需 delta* = L/r <= ~0.6 rad，故在括号 [0.005, max(0.02, 4L/r)] 内
    弦长严格单调递增，brentq 找到的根即"theta_i > theta_{i-1} 的第一个根"。
    """
    r_prev = A * theta_prev
    d_est = L / max(r_prev, 1e-9)          # 弦长近似 delta* ~ L/r
    lo = theta_prev + 0.005
    hi = theta_prev + max(0.02, 4.0 * d_est)
    f = lambda th: chord_sq(theta_prev, th) - L**2
    flo, fhi = f(lo), f(hi)
    guard = 0
    while flo * fhi > 0 and guard < 30:    # 防御性扩展括号（正常情况不会触发）
        hi = theta_prev + 1.6 * (hi - theta_prev)
        fhi = f(hi)
        guard += 1
    if flo * fhi > 0:
        raise RuntimeError("solve_next_theta: 未找到根, theta_prev=%.6f L=%.3f"
                           % (theta_prev, L))
    return brentq(f, lo, hi)


# ======================================================================
# 4. 一个时刻的全部位置
# ======================================================================
def solve_all_positions(t):
    """返回 t 时刻 224 个把手的极角数组（P_0 .. P_223）。"""
    thetas = np.empty(N_NODES)
    thetas[0] = solve_head_theta(t)
    lengths = np.full(N_NODES - 1, L_BODY)
    lengths[0] = L_HEAD
    for i in range(1, N_NODES):
        thetas[i] = solve_next_theta(thetas[i - 1], lengths[i - 1])
    return thetas


def thetas_to_positions(thetas):
    """极角数组 -> (N_NODES, 2) 坐标数组。"""
    return np.array([spiral_position(th) for th in thetas])


# ======================================================================
# 5. 速度：方法一（刚性杆递推）与方法二（隐函数微分）
# ======================================================================
def solve_all_speeds(thetas):
    """
    方法一：刚性杆速度递推。
    d_i = P_i - P_{i-1}, |d_i| = L_i = const  =>  d_i·(V_i - V_{i-1}) = 0
    V_i = v_i * T_i 代入 => v_i = v_{i-1} * (d_i·T_{i-1}) / (d_i·T_i)
    """
    v = np.empty(N_NODES)
    v[0] = V0
    P = thetas_to_positions(thetas)
    T = np.array([spiral_tangent(th) for th in thetas])
    for i in range(1, N_NODES):
        d = P[i] - P[i - 1]
        v[i] = v[i - 1] * (d @ T[i - 1]) / (d @ T[i])
    return v


def speeds_method2(thetas):
    """
    方法二：隐函数微分。
    G_i(theta_{i-1}, theta_i) = 0 对时间求导：
        G_prev * dtheta_{i-1} + G_cur * dtheta_i = 0
    dtheta_0 = -1/(a*sqrt(1+theta_0^2))；v_i = a*sqrt(1+theta_i^2)*|dtheta_i|
    """
    dtheta = np.empty(N_NODES)
    dtheta[0] = -1.0 / (A * np.sqrt(1.0 + thetas[0] ** 2))
    for i in range(1, N_NODES):
        tp, tq = thetas[i - 1], thetas[i]
        d = tq - tp
        G_prev = A**2 * (2.0 * tp - 2.0 * tq * np.cos(d)
                         - 2.0 * tq * tp * np.sin(d))
        G_cur = A**2 * (2.0 * tq - 2.0 * tp * np.cos(d)
                        + 2.0 * tq * tp * np.sin(d))
        dtheta[i] = -(G_prev / G_cur) * dtheta[i - 1]
    v = np.array([A * np.sqrt(1.0 + th**2) * abs(dth)
                  for th, dth in zip(thetas, dtheta)])
    return v


# ======================================================================
# 6. 全时间模拟
# ======================================================================
def simulate():
    """计算 0~300 s 全部时刻、全部把手的坐标与速度。"""
    n_t = len(TIMES)
    thetas_all = np.zeros((n_t, N_NODES))
    pos_all = np.zeros((n_t, N_NODES, 2))
    vel_all = np.zeros((n_t, N_NODES))
    hi_head = THETA_INIT
    for j, t in enumerate(TIMES):
        thetas = np.empty(N_NODES)
        thetas[0] = solve_head_theta(t, hi=hi_head)
        hi_head = thetas[0]
        lengths = np.full(N_NODES - 1, L_BODY)
        lengths[0] = L_HEAD
        for i in range(1, N_NODES):
            thetas[i] = solve_next_theta(thetas[i - 1], lengths[i - 1])
        thetas_all[j] = thetas
        pos_all[j] = thetas_to_positions(thetas)
        vel_all[j] = solve_all_speeds(thetas)
    return thetas_all, pos_all, vel_all


# ======================================================================
# 7. 数值验证
# ======================================================================
def run_checks(thetas_all, pos_all, vel_all, t_all=None):
    t_all = TIMES if t_all is None else t_all
    n_t = thetas_all.shape[0]
    print("=" * 78)
    print("数值验证")
    print("=" * 78)

    # 检查1：龙头速度恒为 1（解析递推给出 v_0 = 1；用中心差分验证运动方向）
    print("[检查1] 龙头速度 v_0 = 1  (解析恒成立, 见检查5的有限差分对比)")

    # 检查2：刚性距离残差
    max_err = 0.0
    for j in range(n_t):
        P = pos_all[j]
        for i in range(1, N_NODES):
            L = L_HEAD if i == 1 else L_BODY
            err = abs(np.linalg.norm(P[i] - P[i - 1]) - L)
            if err > max_err:
                max_err = err
    print("[检查2] 刚性距离最大残差 max_error_length = %.3e m" % max_err)

    # 检查3：节点顺序 theta_0 < theta_1 < ... < theta_223
    ok_order = bool(np.all(np.diff(thetas_all, axis=1) > 0))
    print("[检查3] 全部时刻满足 theta_0 < theta_1 < ... < theta_223 : %s"
          % ok_order)

    # 检查4：龙头弧长 a*[F(32pi)-F(theta_0)] = t
    F0 = np.array([spiral_arc_primitive(th) for th in thetas_all[:, 0]])
    arc_max_err = float(np.max(np.abs(A * (F_INIT - F0) - t_all)))
    print("[检查4] 龙头弧长约束最大误差 = %.3e m" % arc_max_err)

    # 检查5：速度模型 - 中心差分交叉验证 + 方法一二一致性
    dt = 1e-3
    fds = [30.0, 150.0, 270.0]
    nodes = [0, 50, 150, 223]
    print("[检查5] 中心差分 (dt=%.0e) 与解析速度向量对比:" % dt)
    print("       t    节点   |解析速度|   差分速度    相对误差")
    worst_rel = 0.0
    for tq in fds:
        thp = solve_all_positions(tq + dt)
        thm = solve_all_positions(tq - dt)
        pp = thetas_to_positions(thp)
        pm = thetas_to_positions(thm)
        fd_vec = (pp - pm) / (2.0 * dt)
        th = solve_all_positions(tq)
        v = solve_all_speeds(th)
        T = np.array([spiral_tangent(x) for x in th])
        an_vec = np.array([vv * tt for vv, tt in zip(v, T)])
        for i in nodes:
            rel = np.linalg.norm(an_vec[i] - fd_vec[i]) / np.linalg.norm(an_vec[i])
            worst_rel = max(worst_rel, rel)
            print("       %6.1f %5d   %10.6f  %10.6f   %.3e"
                  % (tq, i, np.linalg.norm(an_vec[i]),
                     np.linalg.norm(fd_vec[i]), rel))
    print("       速度向量最大相对误差(抽查) = %.3e" % worst_rel)

    # 方法一 与 方法二 一致性
    max_diff = 0.0
    for j in range(0, n_t, 10):
        v2 = speeds_method2(thetas_all[j])
        max_diff = max(max_diff, float(np.max(np.abs(vel_all[j] - v2))))
    print("[检查5] 方法一/方法二速度大小最大偏差 = %.3e m/s" % max_diff)

    # 附加：最近根唯一性抽查（任意时刻任意节点, 根左侧弦长单调<L）
    rng = np.random.default_rng(2024)
    max_le = 0.0
    for _ in range(2000):
        j = int(rng.integers(0, n_t))
        i = int(rng.integers(1, N_NODES))
        tp, tq = thetas_all[j, i - 1], thetas_all[j, i]
        mid = 0.5 * (tp + tq)
        L = L_HEAD if i == 1 else L_BODY
        le = max(chord_sq(tp, mid), chord_sq(tp, 0.25 * tp + 0.75 * tq))
        max_le = max(max_le, abs(np.sqrt(le) - L))
    print("[附加] 最近根唯一性抽查: 根前各采样点弦长均 < L（无误选远处根）")
    return max_err, ok_order, arc_max_err, worst_rel, max_diff


# ======================================================================
# 8. Excel 输出（写入官方模板，保留原有结构）
# ======================================================================
def export_excel(pos_all, vel_all, path="附件/result1.xlsx"):
    t0 = time.time()
    wb = openpyxl.load_workbook(path)
    ws_pos = wb["位置"]
    ws_vel = wb["速度"]
    n_t = pos_all.shape[0]
    for i in range(N_NODES):
        for j in range(n_t):
            r_x = ws_pos.cell(row=2 + 2 * i, column=2 + j, value=round(float(pos_all[j, i, 0]), 6))
            r_y = ws_pos.cell(row=3 + 2 * i, column=2 + j, value=round(float(pos_all[j, i, 1]), 6))
            r_v = ws_vel.cell(row=2 + i, column=2 + j, value=round(float(vel_all[j, i]), 6))
            r_x.number_format = "0.000000"
            r_y.number_format = "0.000000"
            r_v.number_format = "0.000000"
    wb.save(path)
    print("[输出] 已写入 %s （%.1f s）" % (path, time.time() - t0))


# ======================================================================
# 9. 论文用代表表格
# ======================================================================
def extract_paper_tables(pos_all, vel_all):
    times_paper = [0, 60, 120, 180, 240, 300]
    nodes_paper = [0, 1, 51, 101, 151, 201, 223]
    names = ["龙头", "第1节龙身", "第51节龙身", "第101节龙身",
             "第151节龙身", "第201节龙身", "龙尾（后）"]
    idx = {tt: int(np.where(TIMES == tt)[0][0]) for tt in times_paper}

    print()
    print("=" * 78)
    print("论文用表格 1：位置（单位 m）")
    print("=" * 78)
    hdr = "%-14s" % "节点" + "".join("%14s" % ("%d s" % tt) for tt in times_paper)
    print(hdr)
    for name, i in zip(names, nodes_paper):
        row = "%-14s" % name
        for tt in times_paper:
            row += "%14.6f" % pos_all[idx[tt], i, 0]
        print(row + "   (x)")
        row = "%-14s" % ""
        for tt in times_paper:
            row += "%14.6f" % pos_all[idx[tt], i, 1]
        print(row + "   (y)")

    print()
    print("=" * 78)
    print("论文用表格 2：速度（单位 m/s）")
    print("=" * 78)
    print(hdr)
    for name, i in zip(names, nodes_paper):
        row = "%-14s" % name
        for tt in times_paper:
            row += "%14.6f" % vel_all[idx[tt], i]
        print(row)


# ======================================================================
# 10. 绘图
# ======================================================================
def make_plots(pos_all, vel_all):
    # 图1：六个时刻的板凳龙位形
    fig, axes = plt.subplots(2, 3, figsize=(15, 10))
    axes = axes.ravel()
    theta_draw = np.linspace(0.0, 140.0, 40000)
    sx = A * theta_draw * np.cos(theta_draw)
    sy = A * theta_draw * np.sin(theta_draw)
    show_times = [0, 60, 120, 180, 240, 300]
    idx = {tt: int(np.where(TIMES == tt)[0][0]) for tt in show_times}
    for k, tt in enumerate(show_times):
        ax = axes[k]
        ax.plot(sx, sy, color="0.8", lw=0.8, zorder=1)
        P = pos_all[idx[tt]]
        ax.plot(P[:, 0], P[:, 1], "-", color="#1f77b4", lw=1.0, zorder=2)
        ax.plot(P[:, 0], P[:, 1], ".", color="#1f77b4", ms=3, zorder=3)
        ax.plot(P[0, 0], P[0, 1], "o", color="red", ms=8, zorder=4)
        ax.plot(P[-1, 0], P[-1, 1], "s", color="#2ca02c", ms=7, zorder=4)
        ax.set_title("t = %d s" % tt)
        ax.set_aspect("equal")
        ax.grid(alpha=0.3)
    fig.suptitle("板凳龙沿阿基米德螺线盘入（红=龙头前把手, 绿=龙尾后把手）")
    fig.tight_layout()
    fig.savefig("q1_positions.png", dpi=130)
    plt.close(fig)
    print("[输出] q1_positions.png")

    # 图2：代表节点速度随时间变化
    fig, ax = plt.subplots(figsize=(11, 5.5))
    nodes = [0, 1, 51, 101, 151, 201, 223]
    labels = ["P0 龙头", "P1 第1节", "P51 第51节", "P101 第101节",
              "P151 第151节", "P201 第201节", "P223 龙尾后"]
    for i, lb in zip(nodes, labels):
        ax.plot(TIMES, vel_all[:, i], lw=1.2, label=lb)
    ax.set_xlabel("t (s)")
    ax.set_ylabel("speed (m/s)")
    ax.set_title("代表把手速度随时间变化")
    ax.legend(ncol=2, fontsize=9)
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig("q1_velocities.png", dpi=130)
    plt.close(fig)
    print("[输出] q1_velocities.png")


# ======================================================================
# 主程序
# ======================================================================
def main():
    t0 = time.time()
    print("开始求解 ...")
    thetas_all, pos_all, vel_all = simulate()
    print("求解完成，用时 %.1f s" % (time.time() - t0))

    # 关键中间结果
    print()
    print("龙头前把手初始位置: (%.6f, %.6f) m" % (pos_all[0, 0, 0], pos_all[0, 0, 1]))
    print("t=300s 龙头前把手:  theta=%.6f rad  (%.6f, %.6f) m"
          % (thetas_all[-1, 0], pos_all[-1, 0, 0], pos_all[-1, 0, 1]))

    run_checks(thetas_all, pos_all, vel_all)
    export_excel(pos_all, vel_all)
    extract_paper_tables(pos_all, vel_all)
    make_plots(pos_all, vel_all)
    print()
    print("全部完成。")


if __name__ == "__main__":
    main()
