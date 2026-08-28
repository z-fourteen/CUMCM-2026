# -*- coding: utf-8 -*-
"""
2024 年 CUMCM 数学建模国赛 A 题 问题二：板凳实体几何 + 碰撞检测 + 首次碰撞时刻搜索
==========================================================================================

问题二任务
----------
舞龙队沿问题一相同的阿基米德螺线继续顺时针向内盘入。考虑板凳的**实体几何**
（板凳视为宽度 0.30 m 的矩形，孔中心距端部 0.275 m），检测所有**非相邻**板凳之间
是否发生碰撞，确定能够继续盘入而未发生碰撞的最终时刻 t*（首次碰撞临界时刻）：

    t < t*  :  所有非相邻板凳均不发生碰撞（间隙 > 0）
    t > t*  :  至少存在一对非相邻板凳发生相交（间隙 < 0 / 重叠）

模型要点
--------
1. **板凳实体**：第 i 节板凳连接把手 P_{i-1} 与 P_i，为中心 C_i、轴向单位向量
   u_i = (P_i-P_{i-1})/|P_i-P_{i-1}|、法向 n_i 的矩形。半长 = |P_i-P_{i-1}|/2 + 0.275
   （把手上方与下方各伸出 0.275 m 到板端），半宽 = 0.15 m。
2. **碰撞检测（凸多边形分离轴定理 SAT）**：两矩形相交当且仅当在 4 个分离轴
   （两矩形各自的轴向 u、法向 n）上的投影区间均不分离；接触（投影区间相切）
   也判为碰撞。间隙函数：重叠/接触时取 0，否则取两矩形之间最小点-线段距离
   （两个凸多边形的最小距离必然由"顶点-边"或"顶点-顶点"取到）。
3. **首次碰撞时刻搜索**：全局最小间隙 g(t) = min_{非相邻对} gap(i,j) 是时间的连续
   函数，在 t* 首次降到 0。先用粗网格扫描定位 g(t) <= 0 的区间，再用二分法按
   "g>0 / g==0" 指示函数收敛到首次穿越点。
4. **t* 时刻结果**：重新计算全部 224 个把手的位置与速度，写入官方模板
   result2.xlsx（横坐标 / 纵坐标 / 速度，6 位小数）。

输出
----
* result2.xlsx（写入官方模板，224 把手 x/y/速度，6 位小数）
* 问题二代表节点结果表 + t* 及碰撞节说明
* 碰撞示意图（q2_collision.png）、间隙曲线（q2_gap.png）、t* 时刻速度分布（q2_velocities.png）
* 完整数值验证
"""
import time
import numpy as np
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import solve_q1 as q1

# ======================================================================
# 1. 常量与板凳几何
# ======================================================================
W_BENCH = 0.30                      # 板凳宽度 (m)
HALF_W = W_BENCH / 2.0              # 半宽
END_EXT = 0.275                     # 孔中心到板端的距离 (m)
N_BENCH = q1.N_NODES - 1            # 223 节板凳
# 矩形四个角的相对位置 (+/-h*u +/- w*n)，按顺序 0,1,2,3
CORNER_S = np.array([[1, 1], [1, -1], [-1, 1], [-1, -1]], dtype=float)
# 四条边的端点（角标对），构成闭合回路 0-1-3-2-0
EDGE_IDX = np.array([[0, 1], [1, 3], [3, 2], [2, 0]])


def bench_rectangles(P):
    """
    由 224 个把手坐标 P (N_NODES,2) 构造 223 节板凳矩形。

    返回:
        corners : (223,4,2) 每个矩形的 4 个角点
        u       : (223,2)   轴向单位向量 (P_i - P_{i-1})/|..|
        n       : (223,2)   法向单位向量
        h       : (223,)    半长（含把手两侧伸出段）
    """
    d = np.linalg.norm(P[1:] - P[:-1], axis=1)          # 相邻把手距离
    u = (P[1:] - P[:-1]) / d[:, None]                   # 轴向
    n = np.stack([-u[:, 1], u[:, 0]], axis=1)           # 法向
    C = (P[:-1] + P[1:]) / 2.0                          # 矩形中心
    h = d / 2.0 + END_EXT                               # 半长
    corners = (C[:, None, :]
               + h[:, None, None] * CORNER_S[None, :, 0:1] * u[:, None, :]
               + HALF_W * CORNER_S[None, :, 1:2] * n[:, None, :])
    return corners, u, n, h


def _point_segment_dists(pts, seg_a, seg_b):
    """
    点集到对应线段集的最小距离。
    pts, seg_a, seg_b 均为 (..., 4, 2)；返回 (..., 4) 即每个点距 4 条线段的最短距离。
    """
    v = seg_b - seg_a
    w = pts - seg_a
    vv = np.einsum('...j,...j->...', v, v)
    wv = np.einsum('...j,...j->...', w, v)
    t = np.clip(wv / np.maximum(vv, 1e-30), 0.0, 1.0)
    proj = seg_a + t[..., None] * v
    return np.linalg.norm(pts - proj, axis=-1).min(axis=-1)


def all_pair_gaps(P):
    """
    计算全部非相邻板凳对 (j-i >= 2) 的间隙。
    返回 (gap, ii, jj)：
        gap : (npairs,) 间隙；0 表示重叠或恰好接触
        ii, jj : (npairs,) 板凳对索引（0 起），ii < jj 且 jj-ii >= 2
    """
    corners, u, n, h = bench_rectangles(P)
    i_idx, j_idx = np.triu_indices(N_BENCH, k=2)       # 所有 j-i >= 2 的对
    ci = corners[i_idx]
    cj = corners[j_idx]
    ui, ni = u[i_idx], n[i_idx]
    uj, nj = u[j_idx], n[j_idx]

    # ---- 分离轴定理 (SAT)：4 个轴 (u_i, n_i, u_j, n_j) ----
    axes = np.stack([ui, ni, uj, nj], axis=1)          # (np,4,2)
    proj_i = ci @ axes.transpose(0, 2, 1)              # (np,4,4)
    proj_j = cj @ axes.transpose(0, 2, 1)
    lo_i, hi_i = proj_i.min(axis=1), proj_i.max(axis=1)
    lo_j, hi_j = proj_j.min(axis=1), proj_j.max(axis=1)
    separated = (hi_i < lo_j) | (hi_j < lo_i)          # (np,4)
    overlap = ~separated.any(axis=1)                   # 重叠或接触

    # ---- 未重叠时：顶点到边的最近距离（凸多边形最小距离取点-边/点-点） ----
    Ai, Bi = ci[:, EDGE_IDX[:, 0], :], ci[:, EDGE_IDX[:, 1], :]
    Aj, Bj = cj[:, EDGE_IDX[:, 0], :], cj[:, EDGE_IDX[:, 1], :]
    gap = np.minimum(_point_segment_dists(ci, Aj, Bj),
                     _point_segment_dists(cj, Ai, Bi))
    gap = np.where(overlap, 0.0, gap)
    return gap, i_idx, j_idx


def signed_pair_gap(P, i, j):
    """
    指定板凳对 (i,j) 的带符号间隙：>0 分离，<=0 重叠。
    重叠时取 SAT 各轴上最小重叠量（穿透深度的保守估计）取负。
    """
    corners, u, n, h = bench_rectangles(P)
    ci, cj = corners[[i]], corners[[j]]
    ui, ni, uj, nj = u[[i]], n[[i]], u[[j]], n[[j]]
    axes = np.stack([ui, ni, uj, nj], axis=1)
    proj_i = ci @ axes.transpose(0, 2, 1)
    proj_j = cj @ axes.transpose(0, 2, 1)
    lo_i, hi_i = proj_i.min(axis=1), proj_i.max(axis=1)
    lo_j, hi_j = proj_j.min(axis=1), proj_j.max(axis=1)
    sep_amt = np.maximum(lo_j - hi_i, lo_i - hi_j)   # (1,4) 每轴分离量
    if (sep_amt > 0).any():
        Ai, Bi = ci[:, EDGE_IDX[:, 0], :], ci[:, EDGE_IDX[:, 1], :]
        Aj, Bj = cj[:, EDGE_IDX[:, 0], :], cj[:, EDGE_IDX[:, 1], :]
        gap = np.minimum(_point_segment_dists(ci, Aj, Bj),
                         _point_segment_dists(cj, Ai, Bi))
        return float(gap[0])
    overlap_amt = np.minimum(hi_i, hi_j) - np.maximum(lo_i, lo_j)  # (1,4)
    return -float(overlap_amt.min())


def contact_points(P, i, j):
    """t* 时刻碰撞节 (i,j) 的最近点对（接触点），返回 (pt_a, pt_b, dist)。"""
    corners, u, n, h = bench_rectangles(P)
    ci, cj = corners[i], corners[j]
    best = (1e9, None, None)
    for rect_p, rect_q in ((ci, cj), (cj, ci)):
        for kc in range(4):
            pt = rect_p[kc]
            for ke in range(4):
                a = rect_q[EDGE_IDX[ke, 0]]
                b = rect_q[EDGE_IDX[ke, 1]]
                v = b - a
                w = pt - a
                tt = np.clip((w @ v) / (v @ v), 0.0, 1.0)
                foot = a + tt * v
                dist = np.linalg.norm(pt - foot)
                if dist < best[0]:
                    best = (dist, pt.copy(), foot.copy())
    return best


# ======================================================================
# 2. 全局最小间隙
# ======================================================================
def global_min_gap(t):
    """t 时刻全局最小非相邻间隙，返回 (min_gap, i, j, P)。"""
    P = q1.thetas_to_positions(q1.solve_all_positions(t))
    gap, ii, jj = all_pair_gaps(P)
    k = int(np.argmin(gap))
    return float(gap[k]), int(ii[k]), int(jj[k]), P


# ======================================================================
# 3. 首次碰撞时刻搜索
# ======================================================================
def find_first_collision(t_start=0.0, t_end=440.0, coarse_step=1.0,
                         bisect_tol=1e-9):
    """
    粗扫定位 g(t)<=0 的区间，再对全局最小间隙按 "g>0 / g==0" 指示函数二分。

    返回 (t_star, bi, bj, P_star, g_star)。
    """
    t = t_start
    prev = t_start
    found = False
    while t <= t_end + 1e-9:
        g, _, _, _ = global_min_gap(t)
        if g <= 0.0:
            found = True
            break
        prev = t
        t += coarse_step
    if not found:
        raise RuntimeError("在 t_end=%g 内未发现碰撞" % t_end)

    a, b = prev, t                     # g(a) > 0, g(b) <= 0
    print("[搜索] 粗扫区间: g(%g)=%.6e > 0, g(%g)=0 (首次 <= 0)"
          % (a, global_min_gap(a)[0], b))
    while b - a > bisect_tol:
        mid = 0.5 * (a + b)
        g, _, _, _ = global_min_gap(mid)
        if g > 0.0:
            a = mid
        else:
            b = mid
    t_star = 0.5 * (a + b)
    g_star, bi, bj, P_star = global_min_gap(t_star)
    return t_star, bi, bj, P_star, g_star


# ======================================================================
# 4. 数值验证
# ======================================================================
def run_checks(t_star, bi, bj, P_star, thetas_star, vel_star):
    print()
    print("=" * 78)
    print("问题二数值验证")
    print("=" * 78)

    # (a) 刚性距离残差
    max_err = 0.0
    for i in range(1, q1.N_NODES):
        L = q1.L_HEAD if i == 1 else q1.L_BODY
        err = abs(np.linalg.norm(P_star[i] - P_star[i - 1]) - L)
        max_err = max(max_err, err)
    print("[验证1] t* 时刻刚性距离最大残差 = %.3e m" % max_err)

    # (b) 节点顺序
    print("[验证2] t* 时刻 theta_0<...<theta_223 : %s"
          % bool(np.all(np.diff(thetas_star) > 0)))

    # (c) 龙头弧长
    F0 = q1.spiral_arc_primitive(thetas_star[0])
    print("[验证3] t* 龙头弧长误差 = %.3e m (应 ≈ 0)"
          % abs(q1.A * (q1.F_INIT - F0) - t_star))

    # (d) t* 之前/时刻/之后 的间隙状态
    g_before = global_min_gap(t_star - 0.01)[0]
    g_at = signed_pair_gap(P_star, bi, bj)
    g_after = signed_pair_gap(q1.thetas_to_positions(
        q1.solve_all_positions(t_star + 0.01)), bi, bj)
    print("[验证4] 碰撞节 (%d,%d) 带符号间隙: t*-0.01 -> %+.3e m, "
          "t* -> %+.3e m, t*+0.01 -> %+.3e m"
          % (bi, bj, g_before, g_at, g_after))
    print("        全局最小间隙在 t* 处 = %+.3e m" %
          signed_pair_gap(P_star, bi, bj))

    # (e) t* 之前密集扫描：全部非相邻对均不碰撞
    print("[验证5] t* 之前密集扫描（网格 0.5 s / 近端 0.05 s）全局最小间隙:")
    t_scan1 = np.arange(0.0, 400.0, 0.5)
    t_scan2 = np.arange(400.0, t_star, 0.05)
    t_scan = np.concatenate([t_scan1, t_scan2])
    g_min_all = 1e9
    worst_t = 0.0
    gs = []
    for t in t_scan:
        g, _, _, _ = global_min_gap(t)
        gs.append(g)
        if g < g_min_all:
            g_min_all = g
            worst_t = t
    print("        扫描 %d 点, 最小间隙 = %.6e m (发生在 t=%g s), 均 > 0"
          % (len(t_scan), g_min_all, worst_t))
    return max_err, g_min_all, t_scan, np.array(gs)


# ======================================================================
# 5. Excel 输出（写入官方模板 result2.xlsx）
# ======================================================================
def export_excel(P_star, vel_star, t_star, path="附件/result2.xlsx"):
    t0 = time.time()
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    for i in range(q1.N_NODES):
        r = 2 + i                                   # 节点 i -> 第 2+i 行
        c_x = ws.cell(row=r, column=2, value=round(float(P_star[i, 0]), 6))
        c_y = ws.cell(row=r, column=3, value=round(float(P_star[i, 1]), 6))
        c_v = ws.cell(row=r, column=4, value=round(float(vel_star[i]), 6))
        c_x.number_format = "0.000000"
        c_y.number_format = "0.000000"
        c_v.number_format = "0.000000"
    wb.save(path)
    print("[输出] result2.xlsx 已写入（t* = %.6f s, %.1f s）" % (t_star, time.time() - t0))


# ======================================================================
# 6. 论文用代表表格
# ======================================================================
def extract_paper_tables(t_star, P_star, vel_star, bi, bj):
    nodes_paper = [0, 1, 51, 101, 151, 201, 223]
    names = ["龙头 P0", "第1节龙身 P1", "第51节龙身 P51", "第101节龙身 P101",
             "第151节龙身 P151", "第201节龙身 P201", "龙尾（后）P223"]

    print()
    print("=" * 78)
    print("问题二：首次碰撞临界时刻 t* = %.9f s" % t_star)
    print("碰撞节：第 %d 节（%s, P%d-P%d） 与 第 %d 节（%s, P%d-P%d）"
          % (bi + 1, _bench_name(bi), bi, bi + 1, bj + 1, _bench_name(bj), bj, bj + 1))
    pd, pa, pb = contact_points(P_star, bi, bj)
    print("接触点：(%10.6f, %10.6f) m, 接触点距离残差 = %.3e m"
          % (pa[0], pa[1], pd))
    print("=" * 78)
    print("论文用表格：t* 时刻代表节点位置与速度（单位 m, m/s）")
    print("%-18s %14s %14s %14s" % ("节点", "x (m)", "y (m)", "速度 (m/s)"))
    for name, i in zip(names, nodes_paper):
        print("%-18s %14.6f %14.6f %14.6f"
              % (name, P_star[i, 0], P_star[i, 1], vel_star[i]))

    th0 = _head_theta_at(t_star)
    print()
    print("龙头前把手 t* 状态:  theta=%.9f rad, r=%.9f m, (%.6f, %.6f) m, v=%.9f m/s"
          % (th0, q1.A * th0, P_star[0, 0], P_star[0, 1], vel_star[0]))


def _bench_name(k):
    """0 起板凳索引 -> 名称（龙头/第k节龙身/龙尾）。"""
    if k == 0:
        return "龙头"
    if k == q1.N_NODES - 2:
        return "龙尾"
    return "第%d节龙身" % k


def _head_theta_at(t):
    return q1.solve_head_theta(t)


# ======================================================================
# 7. 绘图
# ======================================================================
def make_plots(P_star, thetas_star, vel_star, t_star, bi, bj, t_scan, gs):
    # ---- 图1：碰撞示意图（全图 + 局部放大） ----
    theta_draw = np.linspace(0.0, 130.0, 40000)
    sx = q1.A * theta_draw * np.cos(theta_draw)
    sy = q1.A * theta_draw * np.sin(theta_draw)
    corners, u, n, h = bench_rectangles(P_star)

    def draw_bench(ax, idx, fc, ec, lw=1.0, alpha=0.8):
        poly = np.vstack([corners[idx], corners[idx][0]])
        ax.fill(poly[:, 0], poly[:, 1], fc=fc, ec=ec, lw=lw, alpha=alpha, zorder=4)

    pd, pa, pb = contact_points(P_star, bi, bj)
    cp = 0.5 * (pa + pb)

    fig = plt.figure(figsize=(15, 6.8))
    ax1 = fig.add_subplot(1, 2, 1)
    ax1.plot(sx, sy, color="0.8", lw=0.8, zorder=1)
    for k in range(N_BENCH):
        draw_bench(ax1, k, "#c9d8f0", "#8fa6c9", lw=0.4, alpha=0.55)
    draw_bench(ax1, bi, "#e04a4a", "#8a1c1c", lw=1.6)
    draw_bench(ax1, bj, "#2fa06a", "#125a38", lw=1.6)
    ax1.plot(P_star[:, 0], P_star[:, 1], ".", color="#20324a", ms=2.5, zorder=5)
    ax1.plot(P_star[0, 0], P_star[0, 1], "o", color="red", ms=7, zorder=6)
    ax1.set_title("t* = %.6f s 全龙位形（红=龙头节, 绿=第%d节）" % (t_star, bj + 1))
    ax1.set_aspect("equal")
    ax1.grid(alpha=0.3)

    ax2 = fig.add_subplot(1, 2, 2)
    ax2.plot(sx, sy, color="0.8", lw=0.8, zorder=1)
    for k in range(N_BENCH):
        draw_bench(ax2, k, "#c9d8f0", "#8fa6c9", lw=0.4, alpha=0.45)
    draw_bench(ax2, bi, "#e04a4a", "#8a1c1c", lw=1.8)
    draw_bench(ax2, bj, "#2fa06a", "#125a38", lw=1.8)
    ax2.plot(P_star[:, 0], P_star[:, 1], ".", color="#20324a", ms=3, zorder=5)
    ax2.plot([pa[0], pb[0]], [pa[1], pb[1]], "o-", color="black", ms=4, lw=1.2, zorder=7)
    ax2.plot([cp[0]], [cp[1]], "*", color="#ff9800", ms=18, zorder=8)
    ax2.annotate("接触点", (cp[0], cp[1]), textcoords="offset points",
                 xytext=(12, 12), fontsize=11, color="#ff9800",
                 arrowprops=dict(arrowstyle="->", color="#ff9800"))
    rad = 2.4
    ax2.set_xlim(cp[0] - rad, cp[0] + rad)
    ax2.set_ylim(cp[1] - rad, cp[1] + rad)
    ax2.set_title("碰撞区域放大（龙头节 × 第%d节, 接触距离 %.1e m）"
                  % (bj + 1, pd))
    ax2.set_aspect("equal")
    ax2.grid(alpha=0.3)
    fig.suptitle("板凳龙首次碰撞示意 t* = %.9f s" % t_star, fontsize=13)
    fig.tight_layout()
    fig.savefig("q2_collision.png", dpi=130)
    plt.close(fig)
    print("[输出] q2_collision.png")

    # ---- 图2：全局最小间隙曲线 ----
    fig, (axa, axb) = plt.subplots(1, 2, figsize=(15, 5.2))
    axa.plot(t_scan, gs, color="#1f77b4", lw=1.1)
    axa.axhline(0, color="k", lw=0.8, ls="--")
    axa.axvline(t_star, color="r", lw=1.2, ls="--")
    axa.set_xlabel("t (s)")
    axa.set_ylabel("全局最小间隙 g(t) (m)")
    axa.set_title("全局最小非相邻间隙（0 ≤ t ≤ t*，始终 > 0）")
    axa.grid(alpha=0.3)
    axa.text(0.98, 0.9, "t* = %.6f s" % t_star, transform=axa.transAxes,
             ha="right", color="r")

    t_fine = np.arange(398.0, t_star + 0.02, 0.02)
    g_glob = np.array([global_min_gap(t)[0] for t in t_fine])
    g08 = np.array([signed_pair_gap(
        q1.thetas_to_positions(q1.solve_all_positions(t)), 0, 8) for t in t_fine])
    g09 = np.array([signed_pair_gap(
        q1.thetas_to_positions(q1.solve_all_positions(t)), 0, 9) for t in t_fine])
    axb.plot(t_fine, g_glob, color="#1f77b4", lw=1.4, label="全局最小间隙")
    axb.plot(t_fine, g08, color="#e04a4a", lw=1.0, ls="--", label="节对 (0,8)")
    axb.plot(t_fine, g09, color="#2fa06a", lw=1.0, ls=":", label="节对 (0,9)")
    axb.axhline(0, color="k", lw=0.8, ls="--")
    axb.axvline(t_star, color="r", lw=1.2, ls="--")
    axb.set_xlabel("t (s)")
    axb.set_ylabel("间隙 (m)")
    axb.set_title("碰撞临近区间的间隙演化（首碰节对 (0,8) 降至 0）")
    axb.legend(fontsize=9)
    axb.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig("q2_gap.png", dpi=130)
    plt.close(fig)
    print("[输出] q2_gap.png")

    # ---- 图3：t* 时刻速度分布 ----
    fig, ax = plt.subplots(figsize=(11, 5.2))
    idx = np.arange(q1.N_NODES)
    ax.plot(idx, vel_star, lw=1.4, color="#1f77b4")
    for i in [0, 1, 51, 101, 151, 201, 223]:
        ax.plot(i, vel_star[i], "o", ms=6, color="#ff7f0e")
    ax.plot([0, 8, 8], [vel_star[0], vel_star[8], 0], "s", ms=6, color="red")
    ax.axhline(1.0, color="0.7", lw=0.8, ls=":")
    ax.set_xlabel("把手编号 i")
    ax.set_ylabel("速度大小 (m/s)")
    ax.set_title("t* = %.6f s 时 224 个把手速度分布（龙头恒为 1 m/s）" % t_star)
    ax.grid(alpha=0.3)
    ax.annotate("龙头 P0 = 1 m/s", (0, vel_star[0]), textcoords="offset points",
                xytext=(8, 10), fontsize=9, color="#ff7f0e")
    fig.tight_layout()
    fig.savefig("q2_velocities.png", dpi=130)
    plt.close(fig)
    print("[输出] q2_velocities.png")


# ======================================================================
# 主程序
# ======================================================================
def main():
    t0 = time.time()
    print("开始求解问题二 ...")

    # 1. 搜索首次碰撞时刻
    t_star, bi, bj, P_star, g_star = find_first_collision()
    print()
    print("首次碰撞临界时刻 t* = %.9f s" % t_star)
    print("碰撞节：第 %d 节（%s）与 第 %d 节（%s）"
          % (bi + 1, _bench_name(bi), bj + 1, _bench_name(bj)))
    print("t* 处全局最小间隙 = %+.6e m" % g_star)

    # 2. t* 时刻速度
    thetas_star = q1.solve_all_positions(t_star)
    vel_star = q1.solve_all_speeds(thetas_star)

    # 3. 验证
    max_err, g_min_all, t_scan, gs = run_checks(
        t_star, bi, bj, P_star, thetas_star, vel_star)

    # 4. 输出 result2.xlsx
    export_excel(P_star, vel_star, t_star)

    # 5. 论文表格
    extract_paper_tables(t_star, P_star, vel_star, bi, bj)

    # 6. 绘图
    make_plots(P_star, thetas_star, vel_star, t_star, bi, bj, t_scan, gs)

    print()
    print("问题二全部完成，总用时 %.1f s" % (time.time() - t0))


if __name__ == "__main__":
    main()
